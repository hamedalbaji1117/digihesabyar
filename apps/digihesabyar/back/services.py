from __future__ import annotations

from dataclasses import dataclass
from typing import Tuple
from typing import List, Tuple, Dict

import math
import re

from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook

from .models import InvoiceFile, InvoiceRow
from .models import PricingTier
from .services import parse_invoice_excel, calculate_price_for_row_count


# --- تبدیل اعداد فارسی/عربی + فرمت‌های متنی به int استاندارد -----------------

PERSIAN_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")
ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")


def _normalize_number(val) -> int:
    """
    هر نوع مقدار عددی از اکسل (int/float/str با ارقام فارسی و کاما و ...) را
    به یک عدد صحیح استاندارد (int) تبدیل می‌کند.
    اگر مقدار نامعتبر باشد، صفر برمی‌گرداند.
    """
    if val is None:
        return 0

    if isinstance(val, (int, float)):
        if isinstance(val, float) and math.isnan(val):
            return 0
        return int(round(val))

    s = str(val)

    # حذف جداکننده هزارگان
    s = s.replace(",", "").replace("٬", "")

    # تبدیل ارقام فارسی و عربی به انگلیسی
    s = s.translate(PERSIAN_DIGITS).translate(ARABIC_DIGITS)

    # حذف کاراکترهای غیر عددی به‌جز علامت منفی در ابتدا
    s = s.strip()
    # فقط علامت منفی در ابتدای رشته را نگه می‌داریم
    is_negative = s.startswith("-")
    s = re.sub(r"[^\d]", "", s)
    if is_negative and s:
        s = "-" + s

    if s in ("", "-", "+"):
        return 0

    try:
        return int(s)
    except ValueError:
        return 0


def normalize_number(val) -> int:
    """
    نسخهٔ public برای استفاده در viewها. فقط wrapper دور _normalize_number است.
    """
    return _normalize_number(val)


# --- ساختار داخلی برای نگه‌داشتن ردیف‌های خوانده شده از اکسل ------------------


@dataclass
class ParsedRow:
    """
    یک ردیف تجمیع شده براساس (نوع فروش، شماره سفارش، DKPC).
    اگر داخل اکسل برای یک سفارش و یک DKPC چند خط باشد، همه روی همین ردیف جمع می‌شوند.
    """

    sale_type: str               # "cash" یا "credit"
    order_id: str
    dkpc: str
    title: str

    sale_amount: int = 0
    is_return: bool = False

    commission_amount: int = 0
    shipping_fee: int = 0
    processing_fee: int = 0
    platform_dev_revenue: int = 0


# --- توابع کمکی برای خواندن شیت‌ها --------------------------------------------


def _get_header_row(ws) -> List[str]:
    """
    اولین ردیفی که حداقل در چند سلول مقدار متنی دارد را به عنوان هدر
    در نظر می‌گیریم و لیست رشته‌ها را برمی‌گردانیم.
    """
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        texts = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if len(texts) >= 3:
            return [str(c).strip() if c is not None else "" for c in row]
    return []


def _find_col(headers: List[str], keywords: List[str]) -> int | None:
    """
    در آرایه هدرها، اولین ستونی که شامل یکی از کلیدواژه‌ها باشد را برمی‌گرداند.
    جستجو case-insensitive است.
    """
    for idx, name in enumerate(headers):
        name_lower = str(name).lower().strip()
        for kw in keywords:
            if kw.lower() in name_lower:
                return idx
    return None


# --- قیمت‌گذاری براساس تعرفه‌ها -------------------------------------------------


def calculate_price_for_row_count(row_count: int) -> int:
    """
    براساس تعداد ردیف‌های فروش، از جدول PricingTier قیمت را پیدا می‌کند.
    اگر هیچ تعرفه‌ای match نشود، صفر برمی‌گرداند.
    """
    if row_count <= 0:
        return 0

    tiers = PricingTier.objects.all().order_by("min_rows")

    for tier in tiers:
        if tier.max_rows is None:
            if tier.min_rows <= row_count:
                return int(tier.price_per_invoice)
        else:
            if tier.min_rows <= row_count <= tier.max_rows:
                return int(tier.price_per_invoice)

    return 0


# --- هسته پردازش اکسل ---------------------------------------------------------


def parse_invoice_excel(file_path: str) -> Tuple[List[ParsedRow], dict, str]:
    """
    فایل اکسل صورتحساب دیجی‌کالا را می‌خواند و:
      - لیست ParsedRow (ردیف‌های تجمیع شده) را برمی‌گرداند
      - دیکشنری log_info شامل لاگ‌های متنی از هر شیت
      - نام شیت‌هایی که پیدا شدند و برخی آمار
    """

    wb = load_workbook(filename=file_path, data_only=True)

    try:
        sheet_names = wb.sheetnames

        log_lines: List[str] = []
        parsed_rows: List[ParsedRow] = []

        # برای دسترسی سریع براساس (sale_type, order_id, dkpc)
        rows_map: Dict[Tuple[str, str, str], ParsedRow] = {}

        def get_or_create(
            sale_type: str,
            order_id,
            dkpc,
            title,
            is_return: bool,
        ) -> ParsedRow:
            """
            اگر این سفارش + DKPC قبلاً دیده شده، همان را برمی‌گرداند؛
            در غیر این صورت ردیف جدیدی می‌سازد.
            """
            key = (sale_type, str(order_id), str(dkpc))
            if key not in rows_map:
                row = ParsedRow(
                    sale_type=sale_type,
                    order_id=str(order_id),
                    dkpc=str(dkpc),
                    title=str(title) if title is not None else "",
                    is_return=is_return,
                )
                rows_map[key] = row
                parsed_rows.append(row)
            return rows_map[key]

        # ----------------- خواندن شیت فروش نقدی و برگشت از فروش ---------------------

        def _read_sales_sheet(
            sheet_name: str,
            sale_type: str,
            is_return_sheet: bool = False,
        ):
            if sheet_name not in sheet_names:
                log_lines.append(f"شیت {sheet_name} پیدا نشد؛ از آن عبور می‌کنیم.")
                return

            ws = wb[sheet_name]
            headers = _get_header_row(ws)
            if not headers:
                log_lines.append(f"شیت {sheet_name}: هدر معتبر پیدا نشد.")
                return

            idx_order = _find_col(headers, ["شماره سفارش", "شناسه سفارش"])
            idx_dkpc = _find_col(headers, ["کد تنوع", "dkpc", "DKPC"])
            idx_title = _find_col(headers, ["عنوان تنوع", "عنوان کالا"])
            idx_amount = _find_col(headers, ["مبلغ نهایی", "مبلغ کل", "مبلغ"])

            if idx_order is None or idx_dkpc is None or idx_amount is None:
                log_lines.append(
                    f"شیت {sheet_name}: ستون‌های کلیدی (شماره سفارش/کد تنوع/مبلغ) پیدا نشد."
                )
                return

            count_rows = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                order_id = row[idx_order]
                dkpc = row[idx_dkpc]
                title = row[idx_title] if idx_title is not None else ""
                amount_raw = row[idx_amount]

                if order_id is None or dkpc is None:
                    continue

                amount = _normalize_number(amount_raw)

                # برای شیت‌های برگشت از فروش، مبلغ معمولاً منفی است؛
                # ما is_return را True می‌کنیم و مبلغ فروش را منفی ذخیره می‌کنیم.
                key = (sale_type, str(order_id), str(dkpc))
                existing_parsed = rows_map.get(key)
                is_return = is_return_sheet or (existing_parsed.is_return if existing_parsed else False)

                parsed = get_or_create(
                    sale_type=sale_type,
                    order_id=order_id,
                    dkpc=dkpc,
                    title=title,
                    is_return=is_return,
                )

                # برای شیت "برگشت از فروش"، مبلغ فروش را منفی می‌کنیم
                if is_return_sheet:
                    parsed.sale_amount -= abs(amount)
                else:
                    parsed.sale_amount += amount

                count_rows += 1

            log_lines.append(
                f"شیت {sheet_name}: {count_rows} ردیف فروش/برگشتی خوانده شد."
            )

        # فروش نقدی
        _read_sales_sheet("فروش", sale_type="cash", is_return_sheet=False)
        _read_sales_sheet("برگشت از فروش", sale_type="cash", is_return_sheet=True)

        # فروش اعتباری
        _read_sales_sheet("فروش اعتباری", sale_type="credit", is_return_sheet=False)
        _read_sales_sheet(
            "برگشت از فروش اعتباری", sale_type="credit", is_return_sheet=True
        )

        # ----------------- خواندن شیت‌های هزینه / کمیسیون / توسعه پلتفرم ----------

        def _apply_cost_sheet(
            sheet_name: str,
            sale_type: str,
            kind: str,
            is_return_sheet: bool = False,
        ):
            """
            kind یکی از مقادیر:
              - "commission"
              - "shipping"
              - "processing"
              - "platform_dev"

            sign = +1 برای هزینه عادی
            sign = -1 برای شیت‌های "برگشت از ..." که باید از هزینه کم شوند.
            """
            if sheet_name not in sheet_names:
                log_lines.append(
                    f"شیت {sheet_name} پیدا نشد (kind={kind})؛ از آن عبور می‌کنیم."
                )
                return

            ws = wb[sheet_name]
            headers = _get_header_row(ws)
            if not headers:
                log_lines.append(f"شیت {sheet_name}: هدر معتبر پیدا نشد.")
                return

            idx_order = _find_col(headers, ["شماره سفارش", "شناسه سفارش"])
            idx_dkpc = _find_col(headers, ["کد تنوع", "dkpc", "DKPC"])
            idx_amount = _find_col(headers, ["مبلغ نهایی", "مبلغ کل", "مبلغ"])

            if idx_order is None or idx_dkpc is None or idx_amount is None:
                log_lines.append(
                    f"شیت {sheet_name}: ستون‌های کلیدی (شماره سفارش/کد تنوع/مبلغ) پیدا نشد."
                )
                return

            sign = -1 if is_return_sheet else 1

            count_rows = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                order_id = row[idx_order]
                dkpc = row[idx_dkpc]
                amount_raw = row[idx_amount]

                if order_id is None or dkpc is None:
                    continue

                amount = _normalize_number(amount_raw) * sign

                key = (sale_type, str(order_id), str(dkpc))
                if key not in rows_map:
                    # اگر برای این سفارش/کدتنوع هنوز ردیفی نداریم، یک ردیف جدید به عنوان فروش با مبلغ فروش صفر می‌سازیم
                    parsed = ParsedRow(
                        sale_type=sale_type,
                        order_id=str(order_id),
                        dkpc=str(dkpc),
                        title="",
                        is_return=False,
                    )
                    rows_map[key] = parsed
                    parsed_rows.append(parsed)
                else:
                    parsed = rows_map[key]

                if kind == "commission":
                    parsed.commission_amount += amount
                elif kind == "shipping":
                    parsed.shipping_fee += amount
                elif kind == "processing":
                    parsed.processing_fee += amount
                elif kind == "platform_dev":
                    parsed.platform_dev_revenue += amount

                count_rows += 1

            log_lines.append(
                f"شیت {sheet_name}: {count_rows} ردیف هزینه روی ردیف‌های فروش اعمال شد (از {sheet_name})."
            )

        # کمیسیون نقدی / اعتباری و برگشت‌هایشان
        _apply_cost_sheet("کمیسیون فروش", "cash", "commission", is_return_sheet=False)
        _apply_cost_sheet(
            "کمیسیون فروش اعتباری", "credit", "commission", is_return_sheet=False
        )
        _apply_cost_sheet("برگشت از کمیسیون فروش", "cash", "commission", is_return_sheet=True)
        _apply_cost_sheet(
            "برگشت از کمیسیون فروش اعتباری",
            "credit",
            "commission",
            is_return_sheet=True,
        )

        # هزینه ارسال
        _apply_cost_sheet("هزینه ارسال", "cash", "shipping", is_return_sheet=False)
        _apply_cost_sheet(
            "هزینه ارسال اعتباری", "credit", "shipping", is_return_sheet=False
        )
        _apply_cost_sheet("برگشت از هزینه ارسال", "cash", "shipping", is_return_sheet=True)
        _apply_cost_sheet(
            "برگشت از هزینه ارسال اعتباری",
            "credit",
            "shipping",
            is_return_sheet=True,
        )

        # هزینه پردازش
        _apply_cost_sheet("هزینه پردازش", "cash", "processing", is_return_sheet=False)
        _apply_cost_sheet(
            "هزینه پردازش اعتباری", "credit", "processing", is_return_sheet=False
        )
        _apply_cost_sheet(
            "برگشت از هزینه پردازش", "cash", "processing", is_return_sheet=True
        )
        _apply_cost_sheet(
            "برگشت از هزینه پردازش اعتباری",
            "credit",
            "processing",
            is_return_sheet=True,
        )

        # درآمد توسعه پلتفرم + برگشت از آن
        _apply_cost_sheet(
            "درآمد توسعه پلتفرم", "credit", "platform_dev", is_return_sheet=False
        )
        _apply_cost_sheet(
            "برگشت از درآمد توسعه پلتفرم",
            "credit",
            "platform_dev",
            is_return_sheet=True,
        )

        # هزینه برگشت از مشتری (به‌عنوان هزینه اضافی روی ردیف فروش)
        _apply_cost_sheet(
            "هزینه برگشت از مشتری", "cash", "shipping", is_return_sheet=False
        )

        # TODO: در آینده شیت‌های دیگر مثل "هزینه ماندگاری در انبار" و ... را هم می‌توان اضافه کرد.

        # در انتها، همه parsed_rows را برمی‌گردانیم
        extra_info = {
            "sheet_names": sheet_names,
            "log": "\n".join(log_lines),
            "row_count": len(parsed_rows),
        }
        return parsed_rows, extra_info, "\n".join(log_lines)
    finally:
        wb.close()  # بستن فایل برای جلوگیری از memory leak


def process_invoice_file(invoice: InvoiceFile) -> Tuple[int, str]:
    """
    فایل اکسل مرتبط با این InvoiceFile را پردازش می‌کند:
      - شیت‌ها را می‌خواند
      - ParsedRow تولید می‌کند
      - InvoiceRowها را می‌سازد/به‌روزرسانی می‌کند
      - row_count و processing_price را روی invoice ست می‌کند
    خروجی:
      - تعداد ردیف‌های نهایی
      - متن لاگ
    """
    invoice.status = InvoiceFile.STATUS_PROCESSING
    invoice.error_message = ""
    invoice.save(update_fields=["status", "error_message"])

    try:
        # استفاده از .path برای فایل‌های محلی، یا open() برای storage های دیگر
        file_path = invoice.original_file.path if hasattr(invoice.original_file, "path") else None
        if file_path is None:
            # برای storage های دیگر (مثل S3/MinIO)، فایل را موقتاً روی دیسک می‌نویسیم
            import tempfile

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                for chunk in invoice.original_file.chunks():
                    tmp_file.write(chunk)
                file_path = tmp_file.name

        parsed_rows, extra_info, log_text = parse_invoice_excel(file_path)

        # حذف ردیف‌های قبلی این صورتحساب و ساخت مجدد
        with transaction.atomic():
            InvoiceRow.objects.filter(invoice=invoice).delete()

            row_count = 0
            for pr in parsed_rows:
                # تعیین is_return براساس شیت مرجوعی یا جمع نهایی مبلغ فروش:
                # اگر مجموع فروش منفی یا صفر شد، این ردیف را مرجوعی در نظر می‌گیریم.
                is_return = pr.is_return or pr.sale_amount <= 0

                if is_return:
                    # برای مرجوعی‌ها، همه مقادیر عددی در دیتابیس را صفر نگه می‌داریم؛
                    # فقط فلگ is_return برای گزارش‌دهی استفاده می‌شود.
                    sale_amount = 0
                    commission_amount = 0
                    shipping_fee = 0
                    processing_fee = 0
                    platform_dev_revenue = 0
                    tax_amount = 0
                else:
                    # برای ردیف‌های فروش عادی، هزینه‌ها را همیشه به‌صورت عدد مثبت ذخیره می‌کنیم.
                    sale_amount = max(0, pr.sale_amount)
                    commission_amount = max(0, pr.commission_amount)
                    shipping_fee = max(0, pr.shipping_fee)
                    processing_fee = max(0, pr.processing_fee)
                    platform_dev_revenue = max(0, pr.platform_dev_revenue)

                    # مالیات خدمات: ۱۰٪ از (کمیسیون + هزینه پردازش)
                    tax_amount = (commission_amount + processing_fee) // 10

                InvoiceRow.objects.create(
                    invoice=invoice,
                    sale_type=pr.sale_type,
                    order_id=pr.order_id,
                    dkpc=pr.dkpc,
                    title=pr.title,
                    sale_amount=sale_amount,
                    commission_amount=commission_amount,
                    shipping_fee=shipping_fee,
                    processing_fee=processing_fee,
                    platform_dev_revenue=platform_dev_revenue,
                    tax_amount=tax_amount,
                    is_return=is_return,
                )
                row_count += 1

            invoice.row_count = row_count
            invoice.processing_price = calculate_price_for_row_count(row_count)
            invoice.status = InvoiceFile.STATUS_DONE
            invoice.save(update_fields=["row_count", "processing_price", "status"])

        return invoice.row_count, log_text

    except Exception as e:
        invoice.status = InvoiceFile.STATUS_ERROR
        invoice.error_message = str(e)
        invoice.save(update_fields=["status", "error_message"])
        return 0, f"خطا در پردازش فایل: {e!r}"
